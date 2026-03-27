"""
통합 6모델 KOSPI 백테스트  (unified_backtest.py)
=================================================

ML 4개 + LLM 2개를 동일한 데이터/파이프라인에서 나란히 백테스트합니다.

모델:
  ML  : lgbm / xgb / lstm / transformer
  LLM : gpt   (GPT-4o-mini, OPENAI_API_KEY)
        claude (Claude Sonnet 4.5, ANTHROPIC_API_KEY)

LLM 전략:
  매 리밸런싱 날짜마다 종목별로 다음 정보를 프롬프트에 포함:
    - 최근 1/3/6개월 수익률
    - 20일 변동성
    - 52주 고/저점 대비 현재가 위치
    - KOSPI 상대 강도
    - 거시지표 스냅샷 (VIX, USD/KRW, KOSPI MoM)
  → 0~100점 투자 점수 + 이유 반환 → TopK 포트폴리오 시뮬

실행:
  # ML만 (API 키 불필요)
  python unified_backtest.py --models lgbm xgb lstm transformer \\
      --tickers 005930 000660 005380 035420 105560 \\
      --test_start 2023-01-01 --test_end 2024-12-31

  # GPT 추가 (OPENAI_API_KEY 필요)
  python unified_backtest.py --models lgbm xgb gpt \\
      --tickers 005930 000660 005380 035420 105560 \\
      --test_start 2023-01-01 --test_end 2024-12-31

  # Claude 추가 (ANTHROPIC_API_KEY 필요)
  python unified_backtest.py --models lgbm xgb claude \\
      --tickers 005930 000660 005380 035420 105560

  # 전체 6개 비교
  python unified_backtest.py --models all \\
      --tickers 005930 000660 005380 035420 105560

  # stub (API/패키지 없이 구조 검증)
  python unified_backtest.py --stub

환경변수:
  OPENAI_API_KEY     — GPT 모델용
  ANTHROPIC_API_KEY  — Claude 모델용
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import sys
import time
import warnings
from dataclasses import dataclass, field
from datetime import date, timedelta
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
from scipy.stats import spearmanr
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s: %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

_HERE = str(Path(__file__).resolve().parent)
if _HERE not in sys.path:
    sys.path.insert(0, _HERE)

# qlib_kospi_backtest.py에서 공통 함수 재사용
from qlib_kospi_backtest import (
    BacktestConfig as MLConfig,
    fetch_ohlcv, fetch_benchmark, build_dataset,
    run_stub as ml_stub, run_portfolio_sim, compute_metrics,
    run_lgbm, run_xgb, run_lstm, run_transformer,
    _build_pred_df, rebalance_dates,
    BacktestMetrics,
)


# ══════════════════════════════════════════════════════════════════════════════
# 통합 설정
# ══════════════════════════════════════════════════════════════════════════════

ML_MODELS  = ["lgbm", "xgb", "lstm", "transformer"]
LLM_MODELS = ["gpt", "claude"]
ALL_MODELS = ML_MODELS + LLM_MODELS


@dataclass
class UnifiedConfig:
    models:       list   = field(default_factory=lambda: ["lgbm", "gpt", "claude"])
    tickers:      list   = field(default_factory=lambda: [
        "005930", "000660", "005380", "035420", "105560",
    ])
    train_start:  str    = "2019-01-01"
    train_end:    str    = "2021-12-31"
    valid_start:  str    = "2022-01-01"
    valid_end:    str    = "2022-12-31"
    test_start:   str    = "2023-01-01"
    test_end:     str    = "2024-12-31"
    topk:         int    = 3
    freq:         str    = "monthly"
    init_cash:    float  = 100_000_000
    cost_bps:     float  = 15.0
    out_dir:      str    = "backtest_results"
    stub:         bool   = False
    cache_dir:    str    = "llm_cache"
    # LLM
    gpt_model:    str    = "gpt-4o-mini"
    claude_model: str    = "claude-sonnet-4-5"
    llm_retry:    int    = 3
    llm_delay:    float  = 1.0   # 초, rate-limit 방지
    # ML 하이퍼파라미터
    num_boost_round:       int   = 500
    early_stopping_rounds: int   = 50
    num_leaves:            int   = 63
    learning_rate:         float = 0.05
    feature_fraction:      float = 0.8
    bagging_fraction:      float = 0.8
    max_depth:             int   = 6
    hidden_size:           int   = 64
    num_layers:            int   = 2
    n_epochs:              int   = 50
    lr:                    float = 1e-3
    batch_size:            int   = 256
    early_stop:            int   = 10
    d_model:               int   = 64
    nhead:                 int   = 4
    dropout:               float = 0.2


# ══════════════════════════════════════════════════════════════════════════════
# LLM 캐시 (API 비용 절약)
# ══════════════════════════════════════════════════════════════════════════════

class LLMCache:
    def __init__(self, cache_dir: str, model_name: str):
        self.path = Path(cache_dir) / model_name
        self.path.mkdir(parents=True, exist_ok=True)

    def _key(self, date_str: str, ticker: str) -> Path:
        return self.path / f"{date_str}_{ticker}.json"

    def get(self, date_str: str, ticker: str) -> Optional[float]:
        p = self._key(date_str, ticker)
        if p.exists():
            try:
                d = json.loads(p.read_text(encoding="utf-8"))
                return float(d["score"])
            except Exception:
                pass
        return None

    def set(self, date_str: str, ticker: str, score: float, reason: str = ""):
        self._key(date_str, ticker).write_text(
            json.dumps({"score": score, "reason": reason},
                       ensure_ascii=False, indent=2),
            encoding="utf-8",
        )


# ══════════════════════════════════════════════════════════════════════════════
# LLM 점수 계산 — 공통 프롬프트 빌더
# ══════════════════════════════════════════════════════════════════════════════

TICKER_NAMES = {
    "005930": "삼성전자", "000660": "SK하이닉스", "005380": "현대차",
    "035420": "NAVER",   "373220": "LG에너지솔루션", "105560": "KB금융",
    "051910": "LG화학",  "207940": "삼성바이오로직스",
    "000270": "기아",    "005490": "POSCO홀딩스",
    "012330": "현대모비스","028260": "삼성물산",
}


def build_llm_prompt(
    ticker: str,
    as_of: str,
    price_series: pd.Series,
    kospi_series: Optional[pd.Series] = None,
    macro_snap: Optional[dict] = None,
) -> tuple[str, str]:
    """
    종목 + 날짜 → (system_prompt, user_prompt)
    """
    name = TICKER_NAMES.get(ticker, ticker)
    p = price_series.dropna()
    p = p[p.index <= pd.Timestamp(as_of)]

    if p.empty:
        return "", ""

    cur = float(p.iloc[-1])

    def ret(n):
        if len(p) < n: return 0.0
        return float(p.iloc[-1] / p.iloc[-n] - 1) * 100

    r1m  = ret(22)
    r3m  = ret(66)
    r6m  = ret(126)
    vol  = float(p.pct_change().iloc[-20:].std() * np.sqrt(252) * 100) if len(p) >= 20 else 0.0
    hi52 = float(p.iloc[-min(252, len(p)):].max())
    lo52 = float(p.iloc[-min(252, len(p)):].min())
    pos52= (cur - lo52) / (hi52 - lo52 + 1e-8) * 100

    # KOSPI 상대강도
    rel = ""
    if kospi_series is not None:
        kp = kospi_series.dropna()
        kp = kp[kp.index <= pd.Timestamp(as_of)]
        if len(kp) >= 22:
            k1m = float(kp.iloc[-1] / kp.iloc[-22] - 1) * 100
            rel = f"\n  - KOSPI 1M 수익률: {k1m:+.1f}%  →  상대강도: {r1m - k1m:+.1f}%p"

    # 거시 스냅샷
    macro_txt = ""
    if macro_snap:
        vix    = macro_snap.get("vix", "N/A")
        usdkrw = macro_snap.get("usdkrw", "N/A")
        kospi  = macro_snap.get("kospi_mom", "N/A")
        macro_txt = (
            f"\n\n거시 환경 스냅샷 ({as_of}):"
            f"\n  - VIX: {vix}"
            f"\n  - USD/KRW: {usdkrw}"
            f"\n  - KOSPI 전월비: {kospi}"
        )

    system = (
        "당신은 KOSPI 주식 전문 퀀트 애널리스트입니다.\n"
        "주어진 종목의 기술적 데이터를 분석하여 향후 1개월 투자 점수(0~100점)를 산출하세요.\n\n"
        "점수 기준:\n"
        "  80~100: 강한 매수 — 상승 모멘텀 뚜렷, 리스크 낮음\n"
        "  60~79 : 매수 고려 — 긍정적 시그널 우세\n"
        "  40~59 : 중립   — 방향성 불명확\n"
        "  20~39 : 매도 고려 — 하락 모멘텀, 리스크 상승\n"
        "  0~19  : 강한 매도 — 명확한 하락 신호\n\n"
        "반드시 아래 JSON 형식으로만 응답하세요:\n"
        '{"score": <0-100 정수>, "reason": "<50자 이내 한국어 근거>"}'
    )

    user = (
        f"종목: {name} ({ticker})\n"
        f"기준일: {as_of}\n\n"
        f"기술적 지표:\n"
        f"  - 현재가: {cur:,.0f}원\n"
        f"  - 1M 수익률: {r1m:+.1f}%\n"
        f"  - 3M 수익률: {r3m:+.1f}%\n"
        f"  - 6M 수익률: {r6m:+.1f}%\n"
        f"  - 20일 변동성(연환산): {vol:.1f}%\n"
        f"  - 52주 고저 위치: {pos52:.1f}% (0%=52주저점, 100%=52주고점)"
        f"{rel}"
        f"{macro_txt}"
    )

    return system, user


# ══════════════════════════════════════════════════════════════════════════════
# GPT 점수 계산
# ══════════════════════════════════════════════════════════════════════════════

def score_with_gpt(
    ticker: str,
    as_of: str,
    price_series: pd.Series,
    kospi_series: Optional[pd.Series],
    macro_snap: Optional[dict],
    cache: LLMCache,
    model: str = "gpt-4o-mini",
    retry: int = 3,
    delay: float = 1.0,
) -> float:
    cached = cache.get(as_of, ticker)
    if cached is not None:
        return cached

    system, user = build_llm_prompt(ticker, as_of, price_series, kospi_series, macro_snap)
    if not system:
        return 50.0

    try:
        from openai import OpenAI
        client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
    except ImportError:
        logger.warning("openai 패키지 없음 — GPT 스킵")
        return 50.0

    for attempt in range(retry):
        try:
            resp = client.chat.completions.create(
                model=model,
                temperature=0.0,
                response_format={"type": "json_object"},
                messages=[
                    {"role": "system", "content": system},
                    {"role": "user",   "content": user},
                ],
            )
            raw = resp.choices[0].message.content
            d   = json.loads(raw)
            score = float(np.clip(d.get("score", 50), 0, 100))
            cache.set(as_of, ticker, score, d.get("reason", ""))
            time.sleep(delay)
            return score
        except Exception as e:
            logger.warning(f"[GPT] {ticker} {as_of} 시도 {attempt+1}/{retry}: {e}")
            time.sleep(delay * (attempt + 1))

    return 50.0


# ══════════════════════════════════════════════════════════════════════════════
# Claude 점수 계산
# ══════════════════════════════════════════════════════════════════════════════

def score_with_claude(
    ticker: str,
    as_of: str,
    price_series: pd.Series,
    kospi_series: Optional[pd.Series],
    macro_snap: Optional[dict],
    cache: LLMCache,
    model: str = "claude-sonnet-4-5",
    retry: int = 3,
    delay: float = 1.0,
) -> float:
    cached = cache.get(as_of, ticker)
    if cached is not None:
        return cached

    system, user = build_llm_prompt(ticker, as_of, price_series, kospi_series, macro_snap)
    if not system:
        return 50.0

    try:
        import anthropic
        client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
    except ImportError:
        logger.warning("anthropic 패키지 없음 — pip install anthropic")
        return 50.0

    for attempt in range(retry):
        try:
            resp = client.messages.create(
                model=model,
                max_tokens=256,
                temperature=0.0,
                system=system,
                messages=[{"role": "user", "content": user}],
            )
            raw = resp.content[0].text.strip()
            # JSON 파싱 — 앞뒤 마크다운 코드블록 제거
            raw = raw.replace("```json", "").replace("```", "").strip()
            d   = json.loads(raw)
            score = float(np.clip(d.get("score", 50), 0, 100))
            cache.set(as_of, ticker, score, d.get("reason", ""))
            time.sleep(delay)
            return score
        except Exception as e:
            logger.warning(f"[Claude] {ticker} {as_of} 시도 {attempt+1}/{retry}: {e}")
            time.sleep(delay * (attempt + 1))

    return 50.0


# ══════════════════════════════════════════════════════════════════════════════
# LLM 시그널 빌드 → pred_df (ML 포맷과 동일)
# ══════════════════════════════════════════════════════════════════════════════

def build_llm_signals(
    provider: str,          # "gpt" | "claude"
    cfg: UnifiedConfig,
    price_df: pd.DataFrame,
    bench_series: Optional[pd.Series],
    ohlcv_dict: dict,
) -> pd.DataFrame:
    """
    리밸런싱 날짜마다 모든 종목에 LLM 점수 계산.
    반환: DataFrame(columns=[date, ticker, score])  ← ML pred_df와 동일 포맷
    """
    reb_dates = rebalance_dates(cfg.test_start, cfg.test_end, cfg.freq)
    model_id  = cfg.gpt_model if provider == "gpt" else cfg.claude_model
    cache     = LLMCache(cfg.cache_dir, f"{provider}_{model_id.replace('/','_')}")
    scorer    = score_with_gpt if provider == "gpt" else score_with_claude

    total  = len(reb_dates) * len(cfg.tickers)
    done   = 0
    records = []

    for reb_ts in reb_dates:
        as_of = str(reb_ts.date())

        # 거시 스냅샷 (KOSPI MoM, VIX 대용)
        macro_snap = {}
        if bench_series is not None and len(bench_series) >= 22:
            kp = bench_series[bench_series.index <= reb_ts]
            if len(kp) >= 22:
                macro_snap["kospi_mom"] = f"{float(kp.iloc[-1]/kp.iloc[-22]-1)*100:+.1f}%"

        for ticker in cfg.tickers:
            done += 1
            if done % 5 == 0:
                logger.info(f"  [{provider.upper()}] 점수 계산 {done}/{total}")

            if ticker not in price_df.columns:
                continue

            price_s = price_df[ticker]

            if cfg.stub:
                score = 50.0 + np.random.normal(0, 15)
            else:
                score = scorer(
                    ticker, as_of, price_s,
                    bench_series, macro_snap if macro_snap else None,
                    cache, model_id,
                    retry=cfg.llm_retry, delay=cfg.llm_delay,
                )

            records.append({
                "date":   pd.Timestamp(as_of),
                "ticker": ticker,
                "score":  float(np.clip(score, 0, 100)),
            })

    df = pd.DataFrame(records)
    logger.info(f"[{provider.upper()}] 시그널 완성: {len(df)}행")
    return df


# ══════════════════════════════════════════════════════════════════════════════
# ML config 변환 헬퍼
# ══════════════════════════════════════════════════════════════════════════════

def to_ml_cfg(cfg: UnifiedConfig) -> MLConfig:
    return MLConfig(
        tickers=cfg.tickers,
        train_start=cfg.train_start, train_end=cfg.train_end,
        valid_start=cfg.valid_start, valid_end=cfg.valid_end,
        test_start=cfg.test_start,   test_end=cfg.test_end,
        topk=cfg.topk, freq=cfg.freq,
        init_cash=cfg.init_cash, cost_bps=cfg.cost_bps,
        out_dir=cfg.out_dir, stub=cfg.stub,
        num_boost_round=cfg.num_boost_round,
        early_stopping_rounds=cfg.early_stopping_rounds,
        num_leaves=cfg.num_leaves, learning_rate=cfg.learning_rate,
        feature_fraction=cfg.feature_fraction, bagging_fraction=cfg.bagging_fraction,
        max_depth=cfg.max_depth, hidden_size=cfg.hidden_size,
        num_layers=cfg.num_layers, n_epochs=cfg.n_epochs,
        lr=cfg.lr, batch_size=cfg.batch_size,
        early_stop=cfg.early_stop, d_model=cfg.d_model,
        nhead=cfg.nhead, dropout=cfg.dropout,
    )


# ══════════════════════════════════════════════════════════════════════════════
# 터미널 출력
# ══════════════════════════════════════════════════════════════════════════════

def print_comparison(results: dict, cfg: UnifiedConfig):
    labels = list(results.keys())
    W      = max(72, 24 + 14 * len(labels))

    print()
    print("=" * W)
    print(f"  {'통합 6모델 KOSPI 백테스트 비교':^{W-4}}")
    print(f"  테스트: {cfg.test_start} ~ {cfg.test_end}  |  TopK={cfg.topk}  |  {cfg.freq}")
    print(f"  종목: {', '.join(cfg.tickers[:6])}{'...' if len(cfg.tickers)>6 else ''}")
    print("=" * W)

    col = 13
    print(f"  {'지표':<24}" + "".join(f"{l:>{col}}" for l in labels))
    print(f"  {'-'*24}" + "".join([f"{'─'*col}"] * len(labels)))

    def row(name, getter, fmt="{:+.2f}"):
        vals = []
        for l in labels:
            try:
                v = getter(results[l])
                vals.append(fmt.format(v))
            except Exception:
                vals.append("  N/A")
        print(f"  {name:<24}" + "".join(f"{v:>{col}}" for v in vals))

    sections = [
        ("📈 수익률", [
            ("ARR (%)",           lambda m: m.annual_return,       "{:+.2f}"),
            ("Total Return (%)",  lambda m: m.total_return,        "{:+.2f}"),
            ("Benchmark (KOSPI)", lambda m: m.benchmark_return,    "{:+.2f}"),
            ("Excess Return (%)", lambda m: m.excess_return,       "{:+.2f}"),
        ]),
        ("⚠️  리스크", [
            ("MDD (%)",           lambda m: m.max_drawdown,        "{:.2f}"),
            ("Annual Vol (%)",    lambda m: m.annual_vol,          "{:.2f}"),
        ]),
        ("⚖️  위험조정", [
            ("Sharpe",            lambda m: m.sharpe,              "{:.4f}"),
            ("Calmar",            lambda m: m.calmar,              "{:.4f}"),
            ("IR",                lambda m: m.information_ratio,   "{:.4f}"),
        ]),
        ("🎯 알파신호", [
            ("IC",                lambda m: m.ic,                  "{:.4f}"),
            ("ICIR",              lambda m: m.icir,                "{:.4f}"),
            ("Rank IC",           lambda m: m.rank_ic,             "{:.4f}"),
            ("Rank ICIR",         lambda m: m.rank_icir,           "{:.4f}"),
        ]),
        ("📊 기타", [
            ("Win Rate (%)",      lambda m: m.win_rate,            "{:.1f}"),
        ]),
    ]

    for sec_name, metrics in sections:
        print(f"\n  ── {sec_name} {'─'*(W-8-len(sec_name))}")
        for name, getter, fmt in metrics:
            row(name, getter, fmt)

    print()
    print("=" * W)

    # 랭킹 요약
    print(f"\n  🏆 모델 랭킹 (ARR 기준)")
    ranked = sorted(labels, key=lambda l: results[l].annual_return, reverse=True)
    for i, l in enumerate(ranked, 1):
        m  = results[l]
        ic_tag = "✅" if m.ic > 0.03 else "⚠️ " if m.ic > 0 else "❌"
        print(f"  {i}위  {l:<14}  ARR={m.annual_return:+.2f}%  "
              f"Sharpe={m.sharpe:.3f}  IC={m.ic:.4f}{ic_tag}  "
              f"MDD={m.max_drawdown:.1f}%")
    print()


# ══════════════════════════════════════════════════════════════════════════════
# Excel 저장
# ══════════════════════════════════════════════════════════════════════════════

C_HDR_BG = "0F172A"; C_HDR_FG = "FFFFFF"
C_SEC_BG = "1E3A5F"; C_SEC_FG = "FFFFFF"
C_ALT    = "F1F5F9"; C_POS    = "16A34A"
C_NEG    = "DC2626"; C_WARN   = "D97706"
C_ML     = "1D4ED8"; C_LLM    = "7C3AED"   # ML=파랑, LLM=보라

def _fill(h): return PatternFill("solid", start_color=h, fgColor=h)
def _font(bold=False, color="1E293B", size=11):
    return Font(bold=bold, color=color, size=size, name="Arial")
def _border():
    s = Side(border_style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)
def _center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _model_color(label: str) -> str:
    return C_LLM if label.upper() in ("GPT", "CLAUDE") else C_ML


def save_excel(results: dict, cfg: UnifiedConfig,
               monthly_navs: dict, out_dir: Path) -> Path:
    wb     = openpyxl.Workbook()
    stamp  = date.today().strftime("%Y%m%d")
    labels = list(results.keys())
    n      = len(labels)

    # ────────────────────────────────────────────────
    # Sheet 1: 지표 비교
    # ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "모델 비교"
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 10
    for i in range(n):
        ws.column_dimensions[get_column_letter(3 + i)].width = 15

    # 타이틀
    ws.row_dimensions[1].height = 38
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=2 + n)
    t = ws.cell(1, 1,
        f"통합 6모델 KOSPI 백테스트 비교  |  {cfg.test_start} ~ {cfg.test_end}")
    t.font      = _font(bold=True, color=C_HDR_FG, size=14)
    t.fill      = _fill(C_HDR_BG)
    t.alignment = _center()

    ws.row_dimensions[2].height = 20
    ws.merge_cells(start_row=2, start_column=1,
                   end_row=2, end_column=2 + n)
    s = ws.cell(2, 1,
        f"TopK={cfg.topk}  |  {cfg.freq}  |  종목: {', '.join(cfg.tickers)}")
    s.font      = _font(color="94A3B8", size=10)
    s.fill      = _fill(C_HDR_BG)
    s.alignment = _center()

    # 컬럼 헤더 (row 3)
    ws.row_dimensions[3].height = 30
    for col, text in enumerate(["지표", "단위"] + labels, 1):
        c = ws.cell(3, col, text)
        if col >= 3:
            lbl = labels[col - 3]
            bg  = C_LLM if lbl.upper() in ("GPT", "CLAUDE") else C_SEC_BG
        else:
            bg = C_SEC_BG
        c.font      = _font(bold=True, color=C_HDR_FG, size=11)
        c.fill      = _fill(bg)
        c.alignment = _center()
        c.border    = _border()

    SECTIONS = [
        ("📈 수익률", [
            ("ARR (%)",             "%",     "annual_return",      True,  "pos"),
            ("Total Return (%)",    "%",     "total_return",       True,  "pos"),
            ("Benchmark KOSPI (%)","% ",     "benchmark_return",   False, "none"),
            ("Excess Return (%)",   "%",     "excess_return",      True,  "pos"),
        ]),
        ("⚠️  리스크", [
            ("MDD (%)",             "%",     "max_drawdown",       True,  "mdd"),
            ("Annual Vol (%)",      "%",     "annual_vol",         False, "none"),
        ]),
        ("⚖️  위험조정", [
            ("Sharpe Ratio",        "×",     "sharpe",             True,  "sharpe"),
            ("Calmar Ratio",        "×",     "calmar",             True,  "pos"),
            ("IR",                  "×",     "information_ratio",  True,  "pos"),
        ]),
        ("🎯 알파신호", [
            ("IC",                  "[-1,1]","ic",                 True,  "ic"),
            ("ICIR",                "×",     "icir",               True,  "pos"),
            ("Rank IC",             "[-1,1]","rank_ic",            True,  "ic"),
            ("Rank ICIR",           "×",     "rank_icir",          True,  "pos"),
        ]),
        ("📊 기타", [
            ("Win Rate (%)",        "%",     "win_rate",           True,  "winrate"),
            ("IC Samples",          "건",    None,                 False, "none"),
        ]),
    ]

    cur = 4
    for sec_name, metrics in SECTIONS:
        ws.row_dimensions[cur].height = 22
        ws.merge_cells(start_row=cur, start_column=1,
                       end_row=cur, end_column=2 + n)
        sc = ws.cell(cur, 1, sec_name)
        sc.font      = _font(bold=True, color=C_SEC_FG, size=10)
        sc.fill      = _fill(C_SEC_BG)
        sc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cur += 1

        for ridx, (metric, unit, attr, colored, ctype) in enumerate(metrics):
            bg = C_ALT if ridx % 2 == 1 else "FFFFFF"
            ws.row_dimensions[cur].height = 22

            nc = ws.cell(cur, 1, metric)
            nc.font      = _font(bold=True, size=10)
            nc.fill      = _fill(bg)
            nc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            nc.border    = _border()

            uc = ws.cell(cur, 2, unit)
            uc.font      = _font(color="64748B", size=9)
            uc.fill      = _fill(bg)
            uc.alignment = _center()
            uc.border    = _border()

            for ci, label in enumerate(labels):
                m   = results[label]
                val = (len(m.ic_monthly) if attr is None else getattr(m, attr))
                vc  = ws.cell(cur, 3 + ci)

                if isinstance(val, int):
                    vc.value = val; vc.number_format = "#,##0"
                elif "%" in unit:
                    vc.value = val / 100
                    vc.number_format = '+0.00%;-0.00%;"-"'
                else:
                    vc.value = val
                    vc.number_format = '+0.0000;-0.0000;"-"'

                # 색상
                if colored:
                    if ctype == "pos":
                        fc = C_POS if val > 0 else C_NEG
                    elif ctype == "mdd":
                        fc = C_NEG if val < -20 else C_WARN if val < -10 else "1E293B"
                    elif ctype == "sharpe":
                        fc = C_POS if val > 1 else C_WARN if val > 0 else C_NEG
                    elif ctype == "ic":
                        fc = C_POS if val > 0.05 else C_WARN if val > 0.02 else C_NEG
                    elif ctype == "winrate":
                        fc = C_POS if val > 55 else C_WARN if val > 45 else C_NEG
                    else:
                        fc = "1E293B"
                else:
                    fc = "1E293B"

                # ML vs LLM 헤더 색
                model_fc = _model_color(label)
                vc.font      = _font(bold=True, color=fc, size=11)
                vc.fill      = _fill(bg)
                vc.alignment = _center()
                vc.border    = _border()

            cur += 1

    # 랭킹 섹션
    cur += 1
    ws.merge_cells(start_row=cur, start_column=1,
                   end_row=cur, end_column=2 + n)
    rh = ws.cell(cur, 1, "🏆 ARR 기준 랭킹")
    rh.font      = _font(bold=True, color=C_SEC_FG, size=10)
    rh.fill      = _fill(C_SEC_BG)
    rh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cur += 1

    ranked = sorted(labels, key=lambda l: results[l].annual_return, reverse=True)
    medals = ["🥇", "🥈", "🥉"] + ["  " for _ in range(20)]
    for ridx, lbl in enumerate(ranked):
        m  = results[lbl]
        bg = C_ALT if ridx % 2 == 1 else "FFFFFF"
        ws.row_dimensions[cur].height = 22
        ws.merge_cells(start_row=cur, start_column=1,
                       end_row=cur, end_column=2 + n)
        rc = ws.cell(cur, 1,
            f"  {medals[ridx]}  {lbl}   ARR={m.annual_return:+.2f}%  "
            f"Sharpe={m.sharpe:.3f}  IC={m.ic:.4f}  MDD={m.max_drawdown:.1f}%")
        model_col = C_LLM if lbl.upper() in ("GPT","CLAUDE") else C_ML
        rc.font      = _font(bold=(ridx == 0), color=model_col, size=10)
        rc.fill      = _fill(bg)
        rc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cur += 1

    ws.freeze_panes = "C4"

    # ────────────────────────────────────────────────
    # Sheet 2: 월별 NAV 비교
    # ────────────────────────────────────────────────
    ws2 = wb.create_sheet("월별 NAV")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 12
    for i in range(n):
        ws2.column_dimensions[get_column_letter(2 + i)].width = 14

    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + n)
    t2 = ws2.cell(1, 1, "월별 정규화 NAV (기준 100)")
    t2.font = _font(bold=True, color=C_HDR_FG, size=13)
    t2.fill = _fill(C_HDR_BG); t2.alignment = _center()

    ws2.cell(2, 1, "날짜").font = _font(bold=True, color=C_HDR_FG)
    ws2.cell(2, 1).fill        = _fill(C_SEC_BG)
    ws2.cell(2, 1).alignment   = _center()

    for i, lbl in enumerate(labels):
        c = ws2.cell(2, 2 + i, lbl)
        bg = C_LLM if lbl.upper() in ("GPT","CLAUDE") else C_SEC_BG
        c.font = _font(bold=True, color=C_HDR_FG)
        c.fill = _fill(bg); c.alignment = _center()

    nav_series = {}
    for lbl, nav in monthly_navs.items():
        mn   = nav.resample("ME").last()
        norm = (mn / mn.iloc[0] * 100).dropna()
        nav_series[lbl] = norm

    all_dates = sorted(set.union(*[set(s.index) for s in nav_series.values()]))
    for ridx, dt in enumerate(all_dates):
        row = 3 + ridx
        bg  = C_ALT if ridx % 2 == 1 else "FFFFFF"
        dc  = ws2.cell(row, 1, dt.strftime("%Y-%m"))
        dc.font = _font(size=10); dc.fill = _fill(bg)
        dc.alignment = _center(); dc.border = _border()

        for i, lbl in enumerate(labels):
            s   = nav_series.get(lbl, pd.Series(dtype=float))
            val = float(s.get(dt, np.nan))
            vc  = ws2.cell(row, 2 + i, None if np.isnan(val) else round(val, 2))
            vc.number_format = "0.00"
            vc.fill = _fill(bg); vc.alignment = _center(); vc.border = _border()
            if not np.isnan(val):
                vc.font = _font(color=C_POS if val >= 100 else C_NEG, size=10)

    ws2.freeze_panes = "B3"

    # ────────────────────────────────────────────────
    # Sheet 3: 원시 데이터
    # ────────────────────────────────────────────────
    ws3 = wb.create_sheet("원시 데이터")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 26
    for i in range(n):
        ws3.column_dimensions[get_column_letter(2 + i)].width = 16

    ws3.cell(1, 1, "지표").font = _font(bold=True, color=C_HDR_FG)
    ws3.cell(1, 1).fill         = _fill(C_HDR_BG); ws3.cell(1, 1).alignment = _center()
    for i, lbl in enumerate(labels):
        bg = C_LLM if lbl.upper() in ("GPT","CLAUDE") else C_SEC_BG
        c  = ws3.cell(1, 2 + i, lbl)
        c.font = _font(bold=True, color=C_HDR_FG)
        c.fill = _fill(bg); c.alignment = _center()

    RAW = [
        ("ARR (%)", "annual_return"), ("Total Return (%)", "total_return"),
        ("Benchmark KOSPI (%)", "benchmark_return"), ("Excess Return (%)", "excess_return"),
        ("MDD (%)", "max_drawdown"), ("Annual Vol (%)", "annual_vol"),
        ("Sharpe", "sharpe"), ("Calmar", "calmar"), ("IR", "information_ratio"),
        ("IC", "ic"), ("ICIR", "icir"), ("Rank IC", "rank_ic"), ("Rank ICIR", "rank_icir"),
        ("Win Rate (%)", "win_rate"),
    ]
    for ridx, (name, attr) in enumerate(RAW):
        row = 2 + ridx
        bg  = C_ALT if ridx % 2 == 1 else "FFFFFF"
        nc  = ws3.cell(row, 1, name)
        nc.font = _font(size=10); nc.fill = _fill(bg)
        nc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        nc.border = _border()
        for i, lbl in enumerate(labels):
            val = getattr(results[lbl], attr)
            vc  = ws3.cell(row, 2 + i, val)
            vc.number_format = "0.0000"; vc.fill = _fill(bg)
            vc.alignment = _center(); vc.border = _border(); vc.font = _font(size=10)

    ws3.freeze_panes = "B2"

    path = out_dir / f"unified_metrics_{stamp}.xlsx"
    wb.save(str(path))
    print(f"  ✅ Excel 저장: {path}")
    return path


def save_csv(results: dict, cfg: UnifiedConfig, out_dir: Path) -> Path:
    rows = []
    for label, m in results.items():
        rows.append({
            "Model": label, "Type": "LLM" if label.upper() in ("GPT","CLAUDE") else "ML",
            "Test Period": f"{cfg.test_start} ~ {cfg.test_end}",
            "Tickers": " ".join(cfg.tickers), "TopK": cfg.topk,
            "ARR (%)": m.annual_return, "Total Return (%)": m.total_return,
            "Benchmark KOSPI (%)": m.benchmark_return, "Excess Return (%)": m.excess_return,
            "MDD (%)": m.max_drawdown, "Annual Vol (%)": m.annual_vol,
            "Sharpe": m.sharpe, "Calmar": m.calmar, "IR": m.information_ratio,
            "IC": m.ic, "ICIR": m.icir, "Rank IC": m.rank_ic, "Rank ICIR": m.rank_icir,
            "Win Rate (%)": m.win_rate, "IC Samples": len(m.ic_monthly),
        })
    df    = pd.DataFrame(rows)
    stamp = date.today().strftime("%Y%m%d")
    path  = out_dir / f"unified_metrics_{stamp}.csv"
    df.to_csv(path, index=False, encoding="utf-8-sig")
    print(f"  ✅ CSV 저장:   {path}")
    return path


# ══════════════════════════════════════════════════════════════════════════════
# main
# ══════════════════════════════════════════════════════════════════════════════

def main():
    p = argparse.ArgumentParser(
        description="통합 6모델 KOSPI 백테스트 (ML×4 + GPT + Claude)",
        formatter_class=argparse.RawTextHelpFormatter,
        epilog="""
예시:
  # ML만
  python unified_backtest.py --models lgbm xgb \\
      --tickers 005930 000660 005380 \\
      --test_start 2023-01-01 --test_end 2024-12-31

  # GPT + Claude 추가 (API 키 필요)
  python unified_backtest.py --models lgbm xgb gpt claude \\
      --tickers 005930 000660 005380 035420 105560

  # 전체 6모델
  python unified_backtest.py --models all

  # stub 검증
  python unified_backtest.py --stub --models all
        """,
    )
    p.add_argument("--models", nargs="+",
                   default=["lgbm", "xgb", "lstm", "transformer", "gpt", "claude"],
                   help="실행할 모델 목록 (all = 전체 6개)")
    p.add_argument("--tickers", nargs="+",
                   default=["005930","000660","005380","035420","105560"])
    p.add_argument("--train_start", default="2019-01-01")
    p.add_argument("--train_end",   default="2021-12-31")
    p.add_argument("--valid_start", default="2022-01-01")
    p.add_argument("--valid_end",   default="2022-12-31")
    p.add_argument("--test_start",  default="2023-01-01")
    p.add_argument("--test_end",    default="2024-12-31")
    p.add_argument("--topk",   type=int,   default=3)
    p.add_argument("--freq",   choices=["monthly","weekly"], default="monthly")
    p.add_argument("--cash",   type=float, default=100_000_000)
    p.add_argument("--cost",   type=float, default=15.0)
    p.add_argument("--out",    default="backtest_results")
    p.add_argument("--cache",  default="llm_cache")
    p.add_argument("--stub",   action="store_true", help="API 없이 구조 검증")
    # LLM
    p.add_argument("--gpt_model",    default="gpt-4o-mini")
    p.add_argument("--claude_model", default="claude-sonnet-4-5")
    p.add_argument("--llm_delay",    type=float, default=0.5,
                   help="API 호출 간 대기 초 (rate-limit 방지)")
    # GBDT
    p.add_argument("--num_boost_round",       type=int,   default=500)
    p.add_argument("--early_stopping_rounds", type=int,   default=50)
    p.add_argument("--num_leaves",            type=int,   default=63)
    p.add_argument("--learning_rate",         type=float, default=0.05)
    p.add_argument("--feature_fraction",      type=float, default=0.8)
    p.add_argument("--bagging_fraction",      type=float, default=0.8)
    p.add_argument("--max_depth",             type=int,   default=6)
    # DL
    p.add_argument("--hidden_size", type=int,   default=64)
    p.add_argument("--num_layers",  type=int,   default=2)
    p.add_argument("--n_epochs",    type=int,   default=50)
    p.add_argument("--lr",          type=float, default=1e-3)
    p.add_argument("--batch_size",  type=int,   default=256)
    p.add_argument("--early_stop",  type=int,   default=10)
    p.add_argument("--d_model",     type=int,   default=64)
    p.add_argument("--nhead",       type=int,   default=4)
    p.add_argument("--dropout",     type=float, default=0.2)
    args = p.parse_args()

    models_to_run = ALL_MODELS if "all" in args.models else args.models

    cfg = UnifiedConfig(
        models=models_to_run, tickers=args.tickers,
        train_start=args.train_start, train_end=args.train_end,
        valid_start=args.valid_start, valid_end=args.valid_end,
        test_start=args.test_start,   test_end=args.test_end,
        topk=args.topk, freq=args.freq,
        init_cash=args.cash, cost_bps=args.cost,
        out_dir=args.out, stub=args.stub, cache_dir=args.cache,
        gpt_model=args.gpt_model, claude_model=args.claude_model,
        llm_delay=args.llm_delay,
        num_boost_round=args.num_boost_round,
        early_stopping_rounds=args.early_stopping_rounds,
        num_leaves=args.num_leaves, learning_rate=args.learning_rate,
        feature_fraction=args.feature_fraction, bagging_fraction=args.bagging_fraction,
        max_depth=args.max_depth, hidden_size=args.hidden_size,
        num_layers=args.num_layers, n_epochs=args.n_epochs,
        lr=args.lr, batch_size=args.batch_size,
        early_stop=args.early_stop, d_model=args.d_model,
        nhead=args.nhead, dropout=args.dropout,
    )

    out_dir = Path(cfg.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    logger.info(f"모델: {models_to_run}")
    logger.info(f"종목: {cfg.tickers}")
    logger.info(f"테스트: {cfg.test_start} ~ {cfg.test_end}  TopK={cfg.topk}")

    # ── API 키 확인 ──
    if not cfg.stub:
        if "gpt" in models_to_run and not os.getenv("OPENAI_API_KEY"):
            logger.warning("OPENAI_API_KEY 없음 — GPT 스킵됩니다")
            models_to_run = [m for m in models_to_run if m != "gpt"]
        if "claude" in models_to_run and not os.getenv("ANTHROPIC_API_KEY"):
            logger.warning("ANTHROPIC_API_KEY 없음 — Claude 스킵됩니다")
            models_to_run = [m for m in models_to_run if m != "claude"]

    # ── 데이터 준비 ──
    ml_cfg = to_ml_cfg(cfg)

    if cfg.stub:
        pred_dummy, price_df = ml_stub(ml_cfg)
        bench_series = None
        X_tr = X_va = X_te = y_tr = y_va = y_te = None
    else:
        logger.info("yfinance OHLCV 수집...")
        ohlcv_dict   = fetch_ohlcv(
            cfg.tickers, cfg.train_start,
            (date.fromisoformat(cfg.test_end) + timedelta(days=60)).isoformat(),
        )
        bench_series = fetch_benchmark(cfg.test_start, cfg.test_end)
        logger.info("Alpha158 피처 엔지니어링...")
        X_tr, y_tr, X_va, y_va, X_te, y_te, price_df = build_dataset(ohlcv_dict, ml_cfg)

    results      = {}
    monthly_navs = {}
    ML_RUNNERS   = {
        "lgbm": run_lgbm, "xgb": run_xgb,
        "lstm": run_lstm, "transformer": run_transformer,
    }

    for model_name in models_to_run:
        label = model_name.upper()
        print(f"\n{'='*55}")
        print(f"  실행: {label}  ({'LLM' if model_name in LLM_MODELS else 'ML'})")
        print(f"{'='*55}")

        try:
            if cfg.stub:
                pred_df, _price = ml_stub(ml_cfg)
                price_df_use    = price_df
            elif model_name in ML_MODELS:
                pred_df      = ML_RUNNERS[model_name](X_tr, y_tr, X_va, y_va, X_te, y_te, ml_cfg)
                price_df_use = price_df
            else:
                # LLM
                pred_df      = build_llm_signals(
                    model_name, cfg, price_df, bench_series, ohlcv_dict if not cfg.stub else {}
                )
                price_df_use = price_df

        except ImportError as e:
            logger.error(f"[{label}] 패키지 없음: {e}")
            continue
        except Exception as e:
            logger.error(f"[{label}] 실패: {e}")
            import traceback; traceback.print_exc()
            continue

        logger.info(f"포트폴리오 시뮬레이션 [{label}]...")
        nav, trades = run_portfolio_sim(pred_df, price_df_use, ml_cfg)

        if len(nav) < 5:
            logger.error(f"[{label}] NAV 부족")
            continue

        logger.info(f"성과 분석 [{label}]...")
        metrics = compute_metrics(nav, pred_df, price_df_use,
                                  bench_series if not cfg.stub else None)

        results[label]      = metrics
        monthly_navs[label] = nav

        # 단일 모델 요약
        print(f"  ARR={metrics.annual_return:+.2f}%  "
              f"Excess={metrics.excess_return:+.2f}%  "
              f"MDD={metrics.max_drawdown:.1f}%  "
              f"Sharpe={metrics.sharpe:.3f}  "
              f"IC={metrics.ic:.4f}  IR={metrics.information_ratio:.3f}")

    if not results:
        print("\n❌ 실행된 모델 없음")
        sys.exit(1)

    # ── 최종 비교 출력 ──
    print_comparison(results, cfg)

    # ── 파일 저장 ──
    print("\n💾 파일 저장 중...")
    save_csv(results, cfg, out_dir)
    save_excel(results, cfg, monthly_navs, out_dir)

    stamp = date.today().strftime("%Y%m%d")
    print(f"\n📁 결과 폴더: {out_dir.resolve()}/")
    print(f"   unified_metrics_{stamp}.xlsx  ← 모델비교 + 월별NAV + 원시데이터")
    print(f"   unified_metrics_{stamp}.csv")


if __name__ == "__main__":
    main()
