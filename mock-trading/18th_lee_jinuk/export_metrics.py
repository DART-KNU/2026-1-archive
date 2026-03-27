"""
Qlib KOSPI 백테스트 지표 추출기  (export_metrics.py)
=====================================================

qlib_kospi_backtest.py와 동일한 파이프라인을 실행하고
ARR / Excess / MDD / Sharpe / Calmar / IC / IR 등
전체 지표를 터미널 + Excel + CSV 로 출력합니다.

실행:
  # stub 모드 (데이터 없이 파이프라인 검증)
  python export_metrics.py --stub

  # 실제 데이터, 단일 모델
  python export_metrics.py --model lgbm \\
      --tickers 005930 000660 005380 035420 \\
      --test_start 2023-01-01 --test_end 2024-12-31

  # 4개 모델 비교
  python export_metrics.py --model all \\
      --tickers 005930 000660 005380 035420 373220 105560 051910 207940
"""

from __future__ import annotations

import argparse
import json
import os
import sys
import warnings
from datetime import date, timedelta
from pathlib import Path

import numpy as np
import pandas as pd
from scipy.stats import spearmanr
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ── qlib_kospi_backtest.py 같은 폴더에 있어야 함 ──
_HERE = str(Path(__file__).resolve().parent)
if _HERE not in sys.path:
    sys.path.insert(0, _HERE)

from qlib_kospi_backtest import (
    BacktestConfig, fetch_ohlcv, fetch_benchmark,
    build_dataset, run_stub, run_portfolio_sim,
    compute_metrics, print_summary,
    run_lgbm, run_xgb, run_lstm, run_transformer,
    MODEL_RUNNERS, MODELS_ALL,
)


# ═══════════════════════════════════════════════════════════════════
# 터미널 출력 — 풀 지표 테이블
# ═══════════════════════════════════════════════════════════════════

def print_full_metrics(results: dict, cfg: BacktestConfig):
    """
    results: {label: BacktestMetrics}
    """
    W = 72

    # ── 헤더 ──
    print()
    print("=" * W)
    print(f"  {'Qlib KOSPI 백테스트 — 전체 지표 요약':^{W-4}}")
    print(f"  테스트: {cfg.test_start} ~ {cfg.test_end}  |  TopK={cfg.topk}  |  {cfg.freq}")
    print(f"  종목: {', '.join(cfg.tickers[:6])}{'...' if len(cfg.tickers)>6 else ''}")
    print("=" * W)

    labels = list(results.keys())
    col_w  = max(12, (W - 28) // max(len(labels), 1))

    def hdr_row(title):
        print(f"\n  ── {title} {'─'*(W-8-len(title))}")

    def row(metric, *vals, fmt="{}", note=""):
        cells = "".join(f"{fmt.format(v):>{col_w}}" for v in vals)
        print(f"  {metric:<24}{cells}   {note}")

    # 컬럼 헤더
    print(f"  {'지표':<24}" + "".join(f"{l:>{col_w}}" for l in labels))
    print(f"  {'-'*24}" + "".join([f"{'─'*col_w}"] * len(labels)))

    def get(label, attr):
        return getattr(results[label], attr)

    # ── 수익률 ──
    hdr_row("수익률 (Return)")
    row("ARR (%)",
        *[f"{get(l,'annual_return'):+.2f}" for l in labels], fmt="{}")
    row("Total Return (%)",
        *[f"{get(l,'total_return'):+.2f}" for l in labels], fmt="{}")
    row("Benchmark KOSPI (%)",
        *[f"{get(l,'benchmark_return'):+.2f}" for l in labels], fmt="{}")
    row("Excess Return (%)",
        *[f"{get(l,'excess_return'):+.2f}" for l in labels], fmt="{}")

    # ── 리스크 ──
    hdr_row("리스크 (Risk)")
    row("MDD (%)",
        *[f"{get(l,'max_drawdown'):.2f}" for l in labels], fmt="{}")
    row("Annual Vol (%)",
        *[f"{get(l,'annual_vol'):.2f}" for l in labels], fmt="{}")

    # ── 위험조정 수익 ──
    hdr_row("위험조정 수익 (Risk-Adjusted)")
    row("Sharpe Ratio",
        *[f"{get(l,'sharpe'):.4f}" for l in labels], fmt="{}")
    row("Calmar Ratio",
        *[f"{get(l,'calmar'):.4f}" for l in labels], fmt="{}")
    row("Information Ratio (IR)",
        *[f"{get(l,'information_ratio'):.4f}" for l in labels], fmt="{}")

    # ── 알파 신호 ──
    hdr_row("알파 신호 (Alpha Signal)")
    row("IC",
        *[f"{get(l,'ic'):.4f}" for l in labels], fmt="{}")
    row("ICIR",
        *[f"{get(l,'icir'):.4f}" for l in labels], fmt="{}")
    row("Rank IC",
        *[f"{get(l,'rank_ic'):.4f}" for l in labels], fmt="{}")
    row("Rank ICIR",
        *[f"{get(l,'rank_icir'):.4f}" for l in labels], fmt="{}")

    # ── 기타 ──
    hdr_row("기타 (Other)")
    row("Win Rate (%)",
        *[f"{get(l,'win_rate'):.1f}" for l in labels], fmt="{}")
    row("IC 샘플 수",
        *[f"{len(get(l,'ic_monthly'))}" for l in labels], fmt="{}")

    print()
    print("=" * W)

    # ── 퀵 판독 가이드 ──
    print()
    print("  📖 판독 가이드")
    print("  ┌─────────────────────────────────────────────────────────────┐")
    print("  │ ARR > 0%            : 절대 수익              │ MDD < -20% : 주의   │")
    print("  │ Excess Return > 0%  : KOSPI 대비 초과 수익   │ Sharpe > 1 : 우수   │")
    print("  │ IC > 0.05           : 유의미한 알파 신호     │ IR > 0.5   : 안정적 │")
    print("  └─────────────────────────────────────────────────────────────┘")
    print()


# ═══════════════════════════════════════════════════════════════════
# CSV 저장
# ═══════════════════════════════════════════════════════════════════

def save_csv(results: dict, cfg: BacktestConfig, out_dir: Path) -> Path:
    rows = []
    for label, m in results.items():
        rows.append({
            "Model":             label,
            "Test Period":       f"{cfg.test_start} ~ {cfg.test_end}",
            "Tickers":           " ".join(cfg.tickers),
            "TopK":              cfg.topk,
            "Freq":              cfg.freq,
            # 수익률
            "ARR (%)":           m.annual_return,
            "Total Return (%)":  m.total_return,
            "Benchmark KOSPI (%)": m.benchmark_return,
            "Excess Return (%)": m.excess_return,
            # 리스크
            "MDD (%)":           m.max_drawdown,
            "Annual Vol (%)":    m.annual_vol,
            # 위험조정
            "Sharpe":            m.sharpe,
            "Calmar":            m.calmar,
            "IR":                m.information_ratio,
            # 알파
            "IC":                m.ic,
            "ICIR":              m.icir,
            "Rank IC":           m.rank_ic,
            "Rank ICIR":         m.rank_icir,
            # 기타
            "Win Rate (%)":      m.win_rate,
            "IC Samples":        len(m.ic_monthly),
        })

    df = pd.DataFrame(rows)
    stamp = date.today().strftime("%Y%m%d")
    path  = out_dir / f"metrics_summary_{stamp}.csv"
    df.to_csv(path, index=False, encoding="utf-8-sig")
    print(f"  ✅ CSV 저장: {path}")
    return path


# ═══════════════════════════════════════════════════════════════════
# Excel 저장 — 전문 포맷
# ═══════════════════════════════════════════════════════════════════

# ── 색상 팔레트 ──
C_HEADER_BG  = "0F172A"   # 네이비 (헤더)
C_HEADER_FG  = "FFFFFF"
C_SECTION_BG = "1E3A5F"   # 다크블루 (섹션)
C_SECTION_FG = "FFFFFF"
C_ALT_ROW    = "F1F5F9"   # 라이트그레이 (짝수 행)
C_POS        = "16A34A"   # 초록 (양수)
C_NEG        = "DC2626"   # 빨강 (음수)
C_WARN       = "D97706"   # 오렌지 (주의)
C_NEUTRAL    = "1E293B"   # 거의 검정
C_GUIDE_BG   = "EFF6FF"   # 라이트블루 (가이드)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def _font(bold=False, color=C_NEUTRAL, size=11, name="Arial") -> Font:
    return Font(bold=bold, color=color, size=size, name=name)

def _border(style="thin") -> Border:
    s = Side(border_style=style, color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _right() -> Alignment:
    return Alignment(horizontal="right", vertical="center")


def _value_color(val: float, metric: str) -> str:
    """지표별 색상 — 양수/음수/경계 구분."""
    if metric in ("MDD (%)", "Annual Vol (%)"):
        return C_NEG if val < -20 else C_WARN if val < -10 else C_NEUTRAL
    if metric in ("ARR (%)", "Total Return (%)", "Excess Return (%)"):
        return C_POS if val > 0 else C_NEG
    if metric == "Sharpe":
        return C_POS if val > 1.0 else C_WARN if val > 0 else C_NEG
    if metric == "IC":
        return C_POS if val > 0.05 else C_WARN if val > 0.02 else C_NEG
    if metric == "IR":
        return C_POS if val > 0.5 else C_WARN if val > 0 else C_NEG
    if metric == "Win Rate (%)":
        return C_POS if val > 55 else C_WARN if val > 45 else C_NEG
    return C_NEUTRAL


def save_excel(results: dict, cfg: BacktestConfig, out_dir: Path,
               monthly_navs: dict | None = None) -> Path:
    """
    results: {label: BacktestMetrics}
    monthly_navs: {label: pd.Series} — 월별 NAV (옵션, Sheet2용)
    """
    wb    = openpyxl.Workbook()
    stamp = date.today().strftime("%Y%m%d")

    # ════════════════════════════════════════════════
    # Sheet 1: 지표 요약
    # ════════════════════════════════════════════════
    ws = wb.active
    ws.title = "지표 요약"
    ws.sheet_view.showGridLines = False

    labels = list(results.keys())
    n_models = len(labels)

    # ── 열 너비 ──
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 18   # 단위/설명
    for i in range(n_models):
        ws.column_dimensions[get_column_letter(3 + i)].width = 16

    # ── 타이틀 행 ──
    ws.row_dimensions[1].height = 36
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=2 + n_models)
    title_cell = ws.cell(1, 1,
        f"Qlib KOSPI 백테스트 — 전체 지표 요약 | {cfg.test_start} ~ {cfg.test_end}")
    title_cell.font      = _font(bold=True, color=C_HEADER_FG, size=14)
    title_cell.fill      = _fill(C_HEADER_BG)
    title_cell.alignment = _center()

    # ── 서브타이틀 ──
    ws.row_dimensions[2].height = 20
    ws.merge_cells(start_row=2, start_column=1,
                   end_row=2, end_column=2 + n_models)
    sub = ws.cell(2, 1,
        f"TopK={cfg.topk}  |  {cfg.freq}  |  종목: {', '.join(cfg.tickers[:6])}{'...' if len(cfg.tickers)>6 else ''}")
    sub.font      = _font(color="94A3B8", size=10)
    sub.fill      = _fill(C_HEADER_BG)
    sub.alignment = _center()

    # ── 컬럼 헤더 (row 3) ──
    ws.row_dimensions[3].height = 28
    for col, text in enumerate(["지표", "단위/설명"] + labels, 1):
        c = ws.cell(3, col, text)
        c.font      = _font(bold=True, color=C_HEADER_FG, size=11)
        c.fill      = _fill(C_SECTION_BG)
        c.alignment = _center()
        c.border    = _border()

    # ── 지표 정의 ──
    SECTIONS = [
        ("📈 수익률 (Return)", [
            ("ARR (%)",            "%",    "annual_return",      True),
            ("Total Return (%)",   "%",    "total_return",       True),
            ("Benchmark (KOSPI %)","%",    "benchmark_return",   False),
            ("Excess Return (%)",  "%",    "excess_return",      True),
        ]),
        ("⚠️ 리스크 (Risk)", [
            ("MDD (%)",            "%",    "max_drawdown",       True),
            ("Annual Vol (%)",     "%",    "annual_vol",         False),
        ]),
        ("⚖️ 위험조정 수익 (Risk-Adjusted)", [
            ("Sharpe Ratio",       "×",    "sharpe",             True),
            ("Calmar Ratio",       "×",    "calmar",             True),
            ("IR (Info. Ratio)",   "×",    "information_ratio",  True),
        ]),
        ("🎯 알파 신호 (Alpha Signal)", [
            ("IC",                 "[-1,1]","ic",                True),
            ("ICIR",               "×",    "icir",               True),
            ("Rank IC",            "[-1,1]","rank_ic",           True),
            ("Rank ICIR",          "×",    "rank_icir",          True),
        ]),
        ("📊 기타 (Other)", [
            ("Win Rate (%)",       "%",    "win_rate",           True),
            ("IC Samples",         "건",   None,                 False),
        ]),
    ]

    cur_row  = 4
    data_rows = []   # (row, metric_name, values) — 나중에 CSV용

    for sec_name, metrics_list in SECTIONS:
        # 섹션 헤더
        ws.row_dimensions[cur_row].height = 22
        ws.merge_cells(start_row=cur_row, start_column=1,
                       end_row=cur_row, end_column=2 + n_models)
        sc = ws.cell(cur_row, 1, sec_name)
        sc.font      = _font(bold=True, color=C_SECTION_FG, size=10)
        sc.fill      = _fill(C_SECTION_BG)
        sc.alignment = Alignment(horizontal="left", vertical="center",
                                 indent=1)
        cur_row += 1

        for idx, (metric, unit, attr, colored) in enumerate(metrics_list):
            ws.row_dimensions[cur_row].height = 22
            bg = C_ALT_ROW if idx % 2 == 1 else "FFFFFF"

            # A열: 지표명
            a = ws.cell(cur_row, 1, metric)
            a.font      = _font(bold=True, size=10)
            a.fill      = _fill(bg)
            a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            a.border    = _border()

            # B열: 단위
            b = ws.cell(cur_row, 2, unit)
            b.font      = _font(color="64748B", size=9)
            b.fill      = _fill(bg)
            b.alignment = _center()
            b.border    = _border()

            row_vals = []
            for col_idx, label in enumerate(labels):
                m   = results[label]
                val = (len(m.ic_monthly) if attr is None
                       else getattr(m, attr))
                row_vals.append(val)

                c = ws.cell(cur_row, 3 + col_idx)

                # 포맷 결정
                if isinstance(val, int):
                    c.value        = val
                    c.number_format = "#,##0"
                elif "%" in unit:
                    c.value        = val / 100        # Excel %형식
                    c.number_format = '+0.00%;-0.00%;"-"'
                else:
                    c.value        = val
                    c.number_format = '+0.0000;-0.0000;"-"'

                # 색상
                fc = _value_color(val, metric) if colored else C_NEUTRAL
                c.font      = _font(bold=True, color=fc, size=11)
                c.fill      = _fill(bg)
                c.alignment = _center()
                c.border    = _border()

            data_rows.append((cur_row, metric, row_vals))
            cur_row += 1

    # ── 판독 가이드 ──
    cur_row += 1
    guide_items = [
        ("ARR > 0%",          "절대 양수 수익"),
        ("Excess > 0%",       "KOSPI 대비 초과 수익"),
        ("MDD > -20%",        "낙폭 주의 구간"),
        ("Sharpe > 1",        "우수한 위험조정 수익"),
        ("Calmar > 0.5",      "MDD 대비 수익 안정"),
        ("IC > 0.05",         "유의미한 알파 신호"),
        ("IC > 0.03",         "약한 신호"),
        ("ICIR > 0.5",        "안정적 IC"),
        ("IR > 0.5",          "안정적 초과수익"),
        ("Win Rate > 55%",    "월별 승률 우수"),
    ]
    ws.merge_cells(start_row=cur_row, start_column=1,
                   end_row=cur_row, end_column=2 + n_models)
    gh = ws.cell(cur_row, 1, "📖 판독 가이드")
    gh.font      = _font(bold=True, color=C_SECTION_FG, size=10)
    gh.fill      = _fill(C_SECTION_BG)
    gh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cur_row += 1

    for i, (cond, desc) in enumerate(guide_items):
        ws.row_dimensions[cur_row].height = 18
        bg = C_GUIDE_BG if i % 2 == 0 else "FFFFFF"
        ws.merge_cells(start_row=cur_row, start_column=1,
                       end_row=cur_row, end_column=2 + n_models)
        gc = ws.cell(cur_row, 1, f"  {cond:<22} →  {desc}")
        gc.font      = _font(size=9, color="1E40AF")
        gc.fill      = _fill(bg)
        gc.alignment = Alignment(horizontal="left", vertical="center")
        cur_row += 1

    # ── 생성일 ──
    cur_row += 1
    ws.merge_cells(start_row=cur_row, start_column=1,
                   end_row=cur_row, end_column=2 + n_models)
    fc = ws.cell(cur_row, 1,
        f"생성: {date.today()}  |  투자 참고용 — 과거 성과가 미래 수익을 보장하지 않습니다.")
    fc.font      = _font(size=8, color="94A3B8")
    fc.alignment = Alignment(horizontal="right", vertical="center")

    # ── 열 고정 (A-B) ──
    ws.freeze_panes = "C4"

    # ════════════════════════════════════════════════
    # Sheet 2: 월별 NAV (월별 수익률 테이블)
    # ════════════════════════════════════════════════
    if monthly_navs:
        ws2 = wb.create_sheet("월별 NAV")
        ws2.sheet_view.showGridLines = False

        # 타이틀
        ws2.column_dimensions["A"].width = 12
        for i, lbl in enumerate(labels):
            ws2.column_dimensions[get_column_letter(2 + i)].width = 14

        ws2.merge_cells(start_row=1, start_column=1,
                        end_row=1, end_column=1 + n_models)
        t = ws2.cell(1, 1, "월별 NAV (기준 100)")
        t.font      = _font(bold=True, color=C_HEADER_FG, size=13)
        t.fill      = _fill(C_HEADER_BG)
        t.alignment = _center()

        # 헤더
        ws2.cell(2, 1, "날짜").font = _font(bold=True, color=C_HEADER_FG)
        ws2.cell(2, 1).fill        = _fill(C_SECTION_BG)
        ws2.cell(2, 1).alignment   = _center()

        for i, lbl in enumerate(labels):
            c = ws2.cell(2, 2 + i, lbl)
            c.font      = _font(bold=True, color=C_HEADER_FG)
            c.fill      = _fill(C_SECTION_BG)
            c.alignment = _center()

        # 공통 날짜 인덱스
        all_nav_series = {}
        for lbl, nav in monthly_navs.items():
            monthly = nav.resample("ME").last()
            norm    = (monthly / monthly.iloc[0] * 100).dropna()
            all_nav_series[lbl] = norm

        all_dates = sorted(set.union(*[set(s.index) for s in all_nav_series.values()]))

        for r_idx, dt in enumerate(all_dates):
            row = 3 + r_idx
            ws2.row_dimensions[row].height = 18
            bg = C_ALT_ROW if r_idx % 2 == 1 else "FFFFFF"

            dc = ws2.cell(row, 1, dt.strftime("%Y-%m"))
            dc.font      = _font(size=10)
            dc.fill      = _fill(bg)
            dc.alignment = _center()
            dc.border    = _border()

            for i, lbl in enumerate(labels):
                s   = all_nav_series[lbl]
                val = float(s.get(dt, np.nan))
                vc  = ws2.cell(row, 2 + i, None if np.isnan(val) else round(val, 2))
                vc.number_format = "0.00"
                vc.fill          = _fill(bg)
                vc.alignment     = _center()
                vc.border        = _border()
                if not np.isnan(val):
                    vc.font = _font(
                        color=C_POS if val >= 100 else C_NEG, size=10
                    )

        ws2.freeze_panes = "B3"

    # ════════════════════════════════════════════════
    # Sheet 3: 원시 데이터 (CSV 호환)
    # ════════════════════════════════════════════════
    ws3  = wb.create_sheet("원시 데이터")
    ws3.sheet_view.showGridLines = False

    ws3.column_dimensions["A"].width = 28
    for i in range(n_models):
        ws3.column_dimensions[get_column_letter(2 + i)].width = 18

    # 헤더
    ws3.cell(1, 1, "지표").font = _font(bold=True, color=C_HEADER_FG, size=11)
    ws3.cell(1, 1).fill         = _fill(C_HEADER_BG)
    ws3.cell(1, 1).alignment    = _center()
    for i, lbl in enumerate(labels):
        c = ws3.cell(1, 2 + i, lbl)
        c.font      = _font(bold=True, color=C_HEADER_FG, size=11)
        c.fill      = _fill(C_HEADER_BG)
        c.alignment = _center()

    RAW_METRICS = [
        ("ARR (%)",             "annual_return"),
        ("Total Return (%)",    "total_return"),
        ("Benchmark KOSPI (%)","benchmark_return"),
        ("Excess Return (%)",   "excess_return"),
        ("MDD (%)",             "max_drawdown"),
        ("Annual Vol (%)",      "annual_vol"),
        ("Sharpe Ratio",        "sharpe"),
        ("Calmar Ratio",        "calmar"),
        ("IR",                  "information_ratio"),
        ("IC",                  "ic"),
        ("ICIR",                "icir"),
        ("Rank IC",             "rank_ic"),
        ("Rank ICIR",           "rank_icir"),
        ("Win Rate (%)",        "win_rate"),
    ]

    for r_idx, (name, attr) in enumerate(RAW_METRICS):
        row = 2 + r_idx
        bg  = C_ALT_ROW if r_idx % 2 == 1 else "FFFFFF"
        nc  = ws3.cell(row, 1, name)
        nc.font      = _font(size=10)
        nc.fill      = _fill(bg)
        nc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        nc.border    = _border()

        for i, lbl in enumerate(labels):
            val = getattr(results[lbl], attr)
            vc  = ws3.cell(row, 2 + i, val)
            vc.number_format = "0.0000"
            vc.fill          = _fill(bg)
            vc.alignment     = _center()
            vc.border        = _border()
            vc.font          = _font(size=10)

    ws3.freeze_panes = "B2"

    # ── 저장 ──
    path = out_dir / f"metrics_full_{stamp}.xlsx"
    wb.save(str(path))
    print(f"  ✅ Excel 저장: {path}")
    return path


# ═══════════════════════════════════════════════════════════════════
# 메인
# ═══════════════════════════════════════════════════════════════════

def main():
    p = argparse.ArgumentParser(
        description="Qlib KOSPI 백테스트 전체 지표 추출기 (터미널 + Excel + CSV)"
    )
    p.add_argument("--model", choices=["lgbm","xgb","lstm","transformer","all"],
                   default="lgbm")
    p.add_argument("--tickers", nargs="+",
                   default=["005930","000660","005380","035420",
                             "373220","105560","051910","207940"])
    p.add_argument("--train_start", default="2019-01-01")
    p.add_argument("--train_end",   default="2021-12-31")
    p.add_argument("--valid_start", default="2022-01-01")
    p.add_argument("--valid_end",   default="2022-12-31")
    p.add_argument("--test_start",  default="2023-01-01")
    p.add_argument("--test_end",    default="2024-12-31")
    p.add_argument("--topk",   type=int,   default=3)
    p.add_argument("--freq",   choices=["monthly","weekly"], default="monthly")
    p.add_argument("--cash",   type=float, default=100_000_000)
    p.add_argument("--cost",   type=float, default=15.0)
    p.add_argument("--out",    default="backtest_results")
    p.add_argument("--stub",   action="store_true")
    # GBDT
    p.add_argument("--num_boost_round",       type=int,   default=500)
    p.add_argument("--early_stopping_rounds", type=int,   default=50)
    p.add_argument("--num_leaves",            type=int,   default=63)
    p.add_argument("--learning_rate",         type=float, default=0.05)
    p.add_argument("--feature_fraction",      type=float, default=0.8)
    p.add_argument("--bagging_fraction",      type=float, default=0.8)
    p.add_argument("--max_depth",             type=int,   default=6)
    # DL
    p.add_argument("--hidden_size", type=int,   default=64)
    p.add_argument("--num_layers",  type=int,   default=2)
    p.add_argument("--n_epochs",    type=int,   default=50)
    p.add_argument("--lr",          type=float, default=1e-3)
    p.add_argument("--batch_size",  type=int,   default=256)
    p.add_argument("--early_stop",  type=int,   default=10)
    p.add_argument("--d_model",     type=int,   default=64)
    p.add_argument("--nhead",       type=int,   default=4)
    p.add_argument("--dropout",     type=float, default=0.2)
    args = p.parse_args()

    cfg = BacktestConfig(
        model=args.model, tickers=args.tickers,
        train_start=args.train_start, train_end=args.train_end,
        valid_start=args.valid_start, valid_end=args.valid_end,
        test_start=args.test_start,   test_end=args.test_end,
        topk=args.topk, freq=args.freq,
        init_cash=args.cash, cost_bps=args.cost,
        out_dir=args.out, stub=args.stub,
        num_boost_round=args.num_boost_round,
        early_stopping_rounds=args.early_stopping_rounds,
        num_leaves=args.num_leaves,
        learning_rate=args.learning_rate,
        feature_fraction=args.feature_fraction,
        bagging_fraction=args.bagging_fraction,
        max_depth=args.max_depth,
        hidden_size=args.hidden_size,
        num_layers=args.num_layers,
        n_epochs=args.n_epochs,
        lr=args.lr,
        batch_size=args.batch_size,
        early_stop=args.early_stop,
        d_model=args.d_model,
        nhead=args.nhead,
        dropout=args.dropout,
    )

    models_to_run = MODELS_ALL if cfg.model == "all" else [cfg.model]
    out_dir = Path(cfg.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    results      = {}   # {label: BacktestMetrics}
    monthly_navs = {}   # {label: pd.Series(NAV)}

    # ── 데이터 준비 ──
    if cfg.stub:
        ohlcv_dict, price_df, bench_series = None, None, None
        X_tr = X_va = X_te = y_tr = y_va = y_te = None
    else:
        print("\n🔄 yfinance OHLCV 수집...")
        ohlcv_dict   = fetch_ohlcv(
            cfg.tickers, cfg.train_start,
            (date.fromisoformat(cfg.test_end) + timedelta(days=60)).isoformat(),
        )
        bench_series = fetch_benchmark(cfg.test_start, cfg.test_end)
        print("🔄 Alpha158 피처 엔지니어링...")
        X_tr, y_tr, X_va, y_va, X_te, y_te, price_df = build_dataset(ohlcv_dict, cfg)

    # ── 모델별 실행 ──
    for model_name in models_to_run:
        label = model_name.upper()
        print(f"\n{'='*55}")
        print(f"  모델 실행: {label}")
        print(f"{'='*55}")

        if cfg.stub:
            pred_df, price_df = run_stub(cfg)
            bench_series      = None
        else:
            runner  = MODEL_RUNNERS[model_name]
            try:
                pred_df = runner(X_tr, y_tr, X_va, y_va, X_te, y_te, cfg)
            except ImportError as e:
                print(f"  ❌ [{label}] 패키지 없음: {e}")
                continue

        print(f"  포트폴리오 시뮬레이션...")
        nav, trades = run_portfolio_sim(pred_df, price_df, cfg)

        if len(nav) < 5:
            print(f"  ❌ [{label}] NAV 생성 실패")
            continue

        print(f"  성과 분석...")
        metrics = compute_metrics(nav, pred_df, price_df, bench_series)

        results[label]      = metrics
        monthly_navs[label] = nav

    if not results:
        print("\n❌ 실행된 모델 없음")
        sys.exit(1)

    # ── 터미널 출력 ──
    print_full_metrics(results, cfg)

    # ── CSV 저장 ──
    print("\n💾 파일 저장 중...")
    csv_path  = save_csv(results, cfg, out_dir)
    xlsx_path = save_excel(results, cfg, out_dir, monthly_navs)

    print(f"\n📁 결과 폴더: {out_dir.resolve()}/")
    print(f"   {xlsx_path.name}  ← Excel (지표요약 + 월별NAV + 원시데이터)")
    print(f"   {csv_path.name}   ← CSV")
    print()


if __name__ == "__main__":
    main()
